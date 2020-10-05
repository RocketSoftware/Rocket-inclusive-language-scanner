#import required libraries
import os, io
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import PySimpleGUI as sg


def find(repo_path, excel_path, msg):
    try:
        #path='C:/GIT'
        #path='C:/Users/phameed/Desktop/Automation/terminology/Test1/Git'
        x=1
        j=0
        z=2
        path=repo_path
        ex_path=excel_path
        wb2 = load_workbook(excel_path)
        sheet1=wb2.sheetnames[0]
        count=wb2['Sheet1'].max_row
        start=time.time()
        exfile=openpyxl.load_workbook(excel_path)
        sheet=exfile.get_sheet_by_name('Sheet1')
        newsheet=exfile.get_sheet_by_name('Sheet2')
        cnt=0
        for path1, subdirs, files in os.walk(path):
            for name in files:
                file_f=os.path.join(path1, name)
                if not file_f.endswith('.dita'): continue
                with open (file_f, "r", encoding = "ISO-8859-1") as f:
                    cnt=cnt+1
                    content=f.readlines()
                    line_count=len(content)
                    for y in range(2, count+1):
                        string=sheet['A' + str(y)].value
                        #string=string1.lower()
                        for item in content:
                            #item1=item.lower()
                            if string in item:
                                try:
                                    newsheet['A' + str(z)].value=file_f
                                    newsheet['B' + str(z)].value=string
                                    newsheet['C' + str(z)].value=sheet['B' + str(y)].value
                                    newsheet['D' + str(z)].value=item
                                    newitem=item.replace(string,sheet['B' + str(y)].value)
                                    newsheet['E' + str(z)].value=newitem
                                except openpyxl.utils.exceptions.IllegalCharacterError:
                                    print(file_f, item)
                                    continue
                                z=z+1
                            else:
                                continue
        end=time.time()
        exec=end-start
        tt=round(exec, 3)
        exfile.save(excel_path)
        msg=str(cnt) + " files are processed at" + str(tt) + " Seconds"
        return(repo_path, excel_path, msg)
    except:
        msg="Not able to save to the terminology sheet. Please close the sheet and try again"
        return(repo_path, excel_path, msg)
def replace(repo_path, excel_path, msg):
    path=repo_path
    wb2 = load_workbook(excel_path)
    sheet1=wb2.sheetnames[0]
    count=wb2['Sheet2'].max_row
    start=time.time()
    exfile=openpyxl.load_workbook(excel_path)
    sheet=exfile.get_sheet_by_name('Sheet1')
    newsheet=exfile.get_sheet_by_name('Sheet2')
    cnt=0
    for y in range(2,count+1):
        res=newsheet['F' + str(y)].value
        if res == 'yes':
            #off_word=newsheet['B' + str(y)].value
            #sugg_word=newsheet['D' + str(y)].value
            line_old=newsheet['D' + str(y)].value
            line_new=newsheet['E' + str(y)].value
            #line=line_i.replace(off_word,sugg_word)
            file_ini=newsheet['A' + str(y)].value
            file_init=file_ini.split('Git')[1]
            file_f=file_init.replace("\\","/")
            file_ff=path+file_f
            try:
                with open (file_ff, "r", encoding = "ISO-8859-1") as f:
                    #print(file_f)
                    content=f.read()
                    f.close()
                    if line_old in content:
                        try:
                            with open (file_ff, "w", encoding = "ISO-8859-1") as fw:
                                content=content.replace(line_old,line_new)
                                fw.write(content)
                                fw.close()
                                log_content1="=================================================="
                                log_content2="ROW" + str(y) + ":  " + "SUCCESS"
                                pre_line="Original Phrase:   " + line_old
                                post_line="Modified Phrase:  " + line_new
                                log_content3="==================================================="
                                log_content=log_content1 + "\n" + log_content2 + "\n" + "File:  " + file_ff + "\n" + pre_line + "\n" + post_line + "\n" + log_content3
                                log_file=open('C:/Users/phameed/Desktop/Automation/terminology/GUI_test/log.txt', 'a+')
                                log_file.write(log_content)
                                cnt=cnt+1
                        except Exception as error:
                                log_content1="==================================================="
                                log_content2="ROW" + str(y) + ":  " + "Failure"
                                log_content12=str(error)
                                log_content3="===================================================="
                                log_content=log_content1 + "\n" + log_content2 + "\n" + file_ff + "\n" + log_content12 + "\n" + log_content3
                                log_file=open('C:/Users/phameed/Desktop/Automation/terminology/GUI_test/log.txt', 'a+')
                                log_file.write(log_content)
                                continue
                    else:
                        log_content1="==================================================="
                        log_content2="ROW" + str(y) + ":  " + "Failure"
                        log_content22="Could not find the original phrase in this document."
                        log_content3="===================================================="
                        log_content=log_content1 + "\n" + log_content2 + "\n" + file_ff + "\n" + log_content22 + "\n" + log_content3
                        log_file=open('C:/Users/phameed/Desktop/Automation/terminology/GUI_test/log.txt', 'a+')
                        log_file.write(log_content)
                        continue
            except Exception as e:
                log_content1="==================================================="
                log_content2="ROW" + str(y) + ":  " + "Failure"
                log_content11=str(e)
                log_content3="===================================================="
                log_content=log_content1 + "\n" + log_content2 + "\n" + file_ff + "\n" + log_content11 + "\n" + log_content3
                log_file=open('C:/Users/phameed/Desktop/Automation/terminology/GUI_test/log.txt', 'a+')
                log_file.write(log_content)
                continue
                
                    
    log_content4="Log Summary" + "\n"
    log_content5=str(cnt) + "  files were modified"
    log_content6= "Terminology sheet used:  " + excel_path
    log_content7="Modified repository path: " + path
    log_pre=log_content4 + "\n" + log_content5 + "\n" + log_content6 + "\n" + log_content7 + "\n" + "\n" + "Detailed Modification Report" + "\n"
    with open('C:/Users/phameed/Desktop/Automation/terminology/GUI_test/log.txt', 'r+') as rp:
        lines=rp.readlines()
        lines.insert(0, log_pre)
        rp.seek(0)
        rp.writelines(lines)
    end=time.time()
    exec=end-start
    tt=round(exec, 3)
    msg=str(cnt) + " files were modified in: " + str(tt) + " seconds"     
    return(repo_path, excel_path, msg)

sg.theme('DarkAmber')	
layout = [  [sg.Text('Find/Replace Deprecated Words')],
            [sg.Text('Enter Documentation Path: '), sg.InputText()],
            [sg.Text('Enter Path For Terminology Sheet: '), sg.InputText()],
            [sg.Button('Find'), sg.Button('Replace')]]
            
window = sg.Window('Terminology Checker', layout)


while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Cancel':	
        break
    if event == 'Find':
        print("button1")
        repo_path=values[0]
        excel_path=values[1]
        msg="string"
        a, b, c=find(repo_path, excel_path, msg)
        sg.Popup(c, keep_on_top=True)
        #break
        #print("the values are: ", values[0], values[1])
        #sg.Popup(str(count) + '   processed in  ' + str(time) + "  seconds", keep_on_top=True)
    if event == 'Replace':
        repo_path=values[0]
        excel_path=values[1]
        msg="empty"
        a, b, c=replace(repo_path, excel_path, msg)
        sg.Popup(c, keep_on_top=True)
        #break
window.close()
